#!/usr/bin/env python3
"""
CRM Query Tools — HubSpot & Salesforce query UI.
AWS session is read from st.session_state (set by app.py).
"""

import io
import json
import time
from datetime import datetime
from typing import Any, Dict, List, Optional

import boto3
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# All outbound HTTP requests use this timeout (seconds).
REQUEST_TIMEOUT = 10

# Delay between each COUNT call in discover mode (seconds).
DISCOVER_DELAY = 0.1

# ── Shared helpers ────────────────────────────────────────────────────────────


def make_excel(records: List[Dict], sheet_name: str = "Results") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not records:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    all_fields = list(dict.fromkeys(k for r in records for k in r.keys()))

    fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")

    for ci, field in enumerate(all_fields, 1):
        cell = ws.cell(row=1, column=ci, value=field)
        cell.fill = fill
        cell.font = font

    for ri, row in enumerate(records, 2):
        for ci, field in enumerate(all_fields, 1):
            val = row.get(field, "")
            ws.cell(row=ri, column=ci, value=str(val) if isinstance(val, dict) else val)

    for col in ws.columns:
        width = min(max((len(str(c.value or "")) for c in col), default=10) + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_download_button(data: bytes, filename: str):
    st.download_button(
        label="⬇️ Download Excel",
        data=data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ── HubSpot client ────────────────────────────────────────────────────────────

HUBSPOT_BASE = "https://api.hubapi.com"
HUBSPOT_STANDARD_OBJECTS = [
    "contacts", "companies", "deals", "tickets",
    "line_items", "products", "quotes", "calls",
    "emails", "meetings", "notes", "tasks", "communications",
]


class HubSpotClient:
    def __init__(self, boto_session: boto3.Session):
        self.sm = boto_session.client("secretsmanager")
        self.token: Optional[str] = None
        self.auth_type: Optional[str] = None

    def load_secret(self, secret_path: str, always_refresh: bool = True):
        resp = self.sm.get_secret_value(SecretId=secret_path)
        creds = json.loads(resp["SecretString"])
        self._authenticate(creds, always_refresh)

    def _authenticate(self, creds: Dict, always_refresh: bool):
        if creds.get("hapikey"):
            self.token = creds["hapikey"]
            self.auth_type = "hapikey"
            return

        has_oauth = all(k in creds for k in ("client_id", "client_secret", "refresh_token"))
        if always_refresh and has_oauth:
            self.token = self._refresh_oauth(creds)
            self.auth_type = "bearer"
            return

        for key in ("access_token", "token", "api_key"):
            if creds.get(key):
                self.token = creds[key]
                self.auth_type = "bearer"
                return

        raise ValueError(f"No usable token found. Keys present: {list(creds.keys())}")

    def _refresh_oauth(self, creds: Dict) -> str:
        r = requests.post(
            "https://api.hubapi.com/oauth/v1/token",
            data={
                "grant_type": "refresh_token",
                "client_id": creds["client_id"],
                "client_secret": creds["client_secret"],
                "refresh_token": creds["refresh_token"],
            },
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        return r.json()["access_token"]

    def _headers(self) -> Dict:
        if self.auth_type == "bearer":
            return {"Authorization": f"Bearer {self.token}", "Content-Type": "application/json"}
        return {"Content-Type": "application/json"}

    def _params(self, extra: Dict = None) -> Dict:
        p = {}
        if self.auth_type == "hapikey":
            p["hapikey"] = self.token
        if extra:
            p.update(extra)
        return p

    def _get(self, path: str, params: Dict = None) -> Dict:
        r = requests.get(
            f"{HUBSPOT_BASE}{path}",
            headers=self._headers(),
            params=self._params(params),
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        return r.json()

    def _post(self, path: str, payload: Dict) -> Dict:
        r = requests.post(
            f"{HUBSPOT_BASE}{path}",
            headers=self._headers(),
            params=self._params(),
            json=payload,
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        return r.json()

    def count(self, obj: str) -> int:
        try:
            d = self._post(f"/crm/v3/objects/{obj}/search", {"filterGroups": [], "limit": 1, "properties": ["hs_object_id"]})
            return d.get("total", 0)
        except Exception:
            return -1

    def get_properties(self, obj: str) -> List[Dict]:
        return self._get(f"/crm/v3/properties/{obj}").get("results", [])

    def list_records(self, obj: str, properties: List[str] = None, limit: int = 20) -> List[Dict]:
        params = {"limit": min(limit, 100)}
        if properties:
            params["properties"] = ",".join(properties)
        return self._get(f"/crm/v3/objects/{obj}", params).get("results", [])

    def search_records(self, obj: str, filters: List[Dict], properties: List[str] = None, limit: int = 100) -> Dict:
        payload: Dict[str, Any] = {
            "filterGroups": [{"filters": filters}] if filters else [],
            "limit": min(limit, 100),
        }
        if properties:
            payload["properties"] = properties
        return self._post(f"/crm/v3/objects/{obj}/search", payload)

    def fetch_all(self, obj: str, prop_names: List[str], limit: int) -> List[Dict]:
        records, after = [], None
        while len(records) < limit:
            batch_size = min(100, limit - len(records))
            payload: Dict[str, Any] = {"filterGroups": [], "properties": prop_names, "limit": batch_size}
            if after:
                payload["after"] = after
            data = self._post(f"/crm/v3/objects/{obj}/search", payload)
            batch = data.get("results", [])
            records.extend(batch)
            after = data.get("paging", {}).get("next", {}).get("after")
            if not after or not batch:
                break
        return records

    def get_schemas(self) -> List[Dict]:
        try:
            return self._get("/crm/v3/schemas").get("results", [])
        except Exception:
            return []

    def flatten(self, record: Dict) -> Dict:
        flat = {"id": record.get("id", "")}
        flat.update(record.get("properties", {}))
        return flat


# ── Salesforce client ─────────────────────────────────────────────────────────


class SalesforceClient:
    def __init__(self, boto_session: boto3.Session):
        self.sm = boto_session.client("secretsmanager")
        self.access_token: Optional[str] = None
        self.instance_url: Optional[str] = None

    def load_secret(self, secret_path: str, always_use_oauth: bool = True):
        resp = self.sm.get_secret_value(SecretId=secret_path)
        creds = json.loads(resp["SecretString"])
        self.instance_url = creds.get("instance_url", "").rstrip("/")
        if not self.instance_url:
            raise ValueError("instance_url not found in credentials")
        if always_use_oauth:
            self.access_token = self._get_token(creds, force=True)
        else:
            self.access_token = creds.get("access_token") or self._get_token(creds)

    def _get_token(self, creds: Dict, force: bool = False) -> str:
        token_url = f"{self.instance_url}/services/oauth2/token"

        if "refresh_token" in creds:
            try:
                r = requests.post(token_url, data={
                    "grant_type": "refresh_token",
                    "client_id": creds["client_id"],
                    "client_secret": creds["client_secret"],
                    "refresh_token": creds["refresh_token"],
                }, timeout=REQUEST_TIMEOUT)
                r.raise_for_status()
                return r.json()["access_token"]
            except Exception:
                pass

        if "username" in creds and "password" in creds:
            try:
                r = requests.post(token_url, data={
                    "grant_type": "password",
                    "client_id": creds["client_id"],
                    "client_secret": creds["client_secret"],
                    "username": creds["username"],
                    "password": creds["password"] + creds.get("security_token", ""),
                }, timeout=REQUEST_TIMEOUT)
                r.raise_for_status()
                return r.json()["access_token"]
            except Exception:
                pass

        r = requests.post(token_url, data={
            "grant_type": "client_credentials",
            "client_id": creds["client_id"],
            "client_secret": creds["client_secret"],
        }, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        return r.json()["access_token"]

    def _headers(self) -> Dict:
        return {"Authorization": f"Bearer {self.access_token}", "Content-Type": "application/json"}

    def query(self, soql: str) -> Dict:
        r = requests.get(
            f"{self.instance_url}/services/data/v59.0/query",
            headers=self._headers(),
            params={"q": soql},
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        return r.json()

    def get_objects(self) -> List[Dict]:
        r = requests.get(
            f"{self.instance_url}/services/data/v59.0/sobjects",
            headers=self._headers(),
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        return r.json().get("sobjects", [])

    def describe(self, sobject: str) -> List[Dict]:
        r = requests.get(
            f"{self.instance_url}/services/data/v59.0/sobjects/{sobject}/describe",
            headers=self._headers(),
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        return [
            {"name": f["name"], "label": f["label"], "type": f["type"], "length": f["length"]}
            for f in r.json().get("fields", [])
        ]

    def count(self, sobject: str) -> int:
        try:
            return self.query(f"SELECT COUNT() FROM {sobject}").get("totalSize", 0)
        except Exception:
            return -1


# ── Main UI ───────────────────────────────────────────────────────────────────

st.title("🔍 CRM Query Tools")

session: Optional[boto3.Session] = st.session_state.get("aws_session")
if not session:
    st.info("👈 Paste your AWS credentials in the sidebar to get started.")
    st.stop()

tab_hs, tab_sf = st.tabs(["🟠 HubSpot", "🔵 Salesforce"])


# ─── HubSpot tab ──────────────────────────────────────────────────────────────

with tab_hs:
    st.header("HubSpot Query Tool")

    hs_customer = st.text_input("Customer name", placeholder="opiniion", key="hs_customer")
    hs_secret = f"{hs_customer.strip()}/hubspot" if hs_customer.strip() else ""
    if hs_secret:
        st.caption(f"Secret path: `{hs_secret}`")
    hs_refresh = st.checkbox("Always refresh OAuth token (recommended)", value=True, key="hs_refresh")

    hs_mode = st.radio("Mode", ["Query objects", "Discover objects"], horizontal=True, key="hs_mode")
    st.divider()

    if hs_mode == "Discover objects":
        hs_filter = st.text_input("Filter by name (leave blank for all)", key="hs_obj_filter")

        if st.button("Search Objects", key="hs_search_btn", type="primary"):
            if not hs_secret:
                st.error("Enter a customer name first.")
            else:
                try:
                    client = HubSpotClient(session)
                    with st.spinner("Fetching credentials from AWS..."):
                        client.load_secret(hs_secret, hs_refresh)

                    with st.spinner("Fetching object schemas..."):
                        schemas = client.get_schemas()

                    all_objects = [{"name": o, "label": o.title(), "type": "standard"} for o in HUBSPOT_STANDARD_OBJECTS]
                    for s in schemas:
                        all_objects.append({
                            "name": s.get("fullyQualifiedName", s.get("name", "")),
                            "label": s.get("labels", {}).get("singular", s.get("name", "")),
                            "type": "custom",
                        })

                    if hs_filter:
                        term = hs_filter.lower()
                        all_objects = [o for o in all_objects if term in o["name"].lower() or term in o["label"].lower()]

                    st.info(f"{len(all_objects)} object(s) found — counting records...")
                    rows = []
                    prog = st.progress(0)
                    status_text = st.empty()
                    for i, obj in enumerate(all_objects):
                        status_text.caption(f"({i + 1}/{len(all_objects)}) {obj['name']}")
                        count = client.count(obj["name"])
                        rows.append({**obj, "record_count": count if isinstance(count, int) and count >= 0 else "Error"})
                        prog.progress((i + 1) / len(all_objects))
                        time.sleep(DISCOVER_DELAY)
                    status_text.empty()

                    st.success("Done")
                    st.dataframe(rows, use_container_width=True)

                except Exception as exc:
                    st.error(f"Error: {exc}")

    else:
        col1, col2 = st.columns(2)
        with col1:
            hs_object = st.text_input("Object type", value="contacts", key="hs_object",
                                      help="e.g. contacts, companies, deals, or a custom object name")
            hs_qtype = st.selectbox("Query type", ["count", "list", "all", "shape", "search"], key="hs_qtype")
        with col2:
            hs_limit = st.number_input("Record limit", min_value=1, max_value=10000, value=100, key="hs_limit")
            hs_props = st.text_input("Properties (comma-separated, blank = default)", key="hs_props")

        hs_filters_parsed = None
        if hs_qtype == "search":
            st.subheader("Search filters")
            st.caption("All filters are ANDed together. Operators: EQ, NEQ, LT, LTE, GT, GTE, CONTAINS_TOKEN, HAS_PROPERTY, etc.")
            hs_filters_text = st.text_area(
                "Filters JSON",
                value='[\n  {"propertyName": "lifecyclestage", "operator": "EQ", "value": "customer"}\n]',
                height=120,
                key="hs_filters",
            )
            try:
                hs_filters_parsed = json.loads(hs_filters_text)
            except Exception:
                st.warning("Invalid filter JSON — fix before running")

        if st.button("Run Query", key="hs_run", type="primary"):
            if not hs_secret:
                st.error("Enter a customer name first.")
            elif not hs_object:
                st.error("Enter an object type.")
            elif hs_qtype == "search" and hs_filters_parsed is None:
                st.error("Fix the filter JSON before running.")
            else:
                try:
                    client = HubSpotClient(session)
                    with st.spinner("Fetching credentials from AWS..."):
                        client.load_secret(hs_secret, hs_refresh)

                    props = [p.strip() for p in hs_props.split(",") if p.strip()] if hs_props else []

                    if hs_qtype == "count":
                        with st.spinner("Counting..."):
                            total = client.count(hs_object)
                        st.metric("Total records", f"{total:,}")

                    elif hs_qtype == "list":
                        with st.spinner("Fetching records..."):
                            records = client.list_records(hs_object, props or None, hs_limit)
                        flat = [client.flatten(r) for r in records]
                        st.success(f"{len(flat)} records returned")
                        st.dataframe(flat, use_container_width=True)

                    elif hs_qtype == "shape":
                        with st.spinner("Fetching properties..."):
                            all_props = client.get_properties(hs_object)
                        rows = [
                            {"name": p["name"], "label": p["label"], "type": p["type"],
                             "fieldType": p["fieldType"], "group": p["groupName"]}
                            for p in all_props
                        ]
                        st.success(f"{len(rows)} properties found")
                        st.dataframe(rows, use_container_width=True)
                        excel_download_button(make_excel(rows, "Object Shape"), f"{hs_object}_shape_{ts()}.xlsx")

                    elif hs_qtype == "all":
                        with st.spinner("Fetching all properties..."):
                            all_props = client.get_properties(hs_object)
                        prop_names = [p["name"] for p in all_props]
                        with st.spinner(f"Fetching up to {hs_limit} records ({len(prop_names)} properties)..."):
                            records = client.fetch_all(hs_object, prop_names, hs_limit)
                        flat = [client.flatten(r) for r in records]
                        st.success(f"{len(flat)} records, {len(prop_names)} properties")
                        st.dataframe(flat, use_container_width=True)
                        excel_download_button(make_excel(flat, "Query Results"), f"{hs_object}_records_{ts()}.xlsx")

                    elif hs_qtype == "search":
                        with st.spinner("Searching..."):
                            result = client.search_records(hs_object, hs_filters_parsed, props or None, hs_limit)
                        records = result.get("results", [])
                        total = result.get("total", len(records))
                        flat = [client.flatten(r) for r in records]
                        st.success(f"{total:,} total matching — {len(flat)} returned")
                        st.dataframe(flat, use_container_width=True)

                except Exception as exc:
                    st.error(f"Error: {exc}")


# ─── Salesforce tab ────────────────────────────────────────────────────────────

with tab_sf:
    st.header("Salesforce Query Tool")

    sf_customer = st.text_input("Customer name", placeholder="conga", key="sf_customer")
    sf_secret = f"{sf_customer.strip()}/salesforce" if sf_customer.strip() else ""
    if sf_secret:
        st.caption(f"Secret path: `{sf_secret}`")
    sf_oauth = st.checkbox("Always use OAuth (recommended)", value=True, key="sf_oauth")

    sf_mode = st.radio("Mode", ["Query objects", "Discover objects"], horizontal=True, key="sf_mode")
    st.divider()

    if sf_mode == "Discover objects":
        sf_filter = st.text_input("Filter by name (leave blank for all)", key="sf_obj_filter")

        if st.button("Search Objects", key="sf_search_btn", type="primary"):
            if not sf_secret:
                st.error("Enter a customer name first.")
            else:
                try:
                    client = SalesforceClient(session)
                    with st.spinner("Fetching credentials from AWS..."):
                        client.load_secret(sf_secret, sf_oauth)

                    with st.spinner("Fetching Salesforce objects..."):
                        all_objects = client.get_objects()

                    if sf_filter:
                        term = sf_filter.lower()
                        all_objects = [
                            o for o in all_objects
                            if term in o["name"].lower() or term in o.get("label", "").lower()
                        ]

                    st.info(f"{len(all_objects)} object(s) found — counting records...")
                    rows = []
                    prog = st.progress(0)
                    status_text = st.empty()
                    for i, obj in enumerate(all_objects):
                        is_queryable = obj.get("queryable", False)
                        status_text.caption(f"({i + 1}/{len(all_objects)}) {obj['name']}")
                        if is_queryable:
                            count = client.count(obj["name"])
                            record_count = count if isinstance(count, int) and count >= 0 else "Error"
                            time.sleep(DISCOVER_DELAY)
                        else:
                            record_count = "N/A"
                        rows.append({
                            "name": obj["name"],
                            "label": obj.get("label", ""),
                            "queryable": is_queryable,
                            "record_count": record_count,
                        })
                        prog.progress((i + 1) / len(all_objects))
                    status_text.empty()

                    st.success("Done")
                    st.dataframe(rows, use_container_width=True)

                except Exception as exc:
                    st.error(f"Error: {exc}")

    else:
        col1, col2 = st.columns(2)
        with col1:
            sf_object = st.text_input("Salesforce object", placeholder="e.g. Contact, Account, Asset", key="sf_object")
            sf_qtype = st.selectbox("Query type", ["count", "list", "all", "shape", "custom"], key="sf_qtype")
        with col2:
            sf_limit = st.number_input("Record limit (max 200 for 'all')", min_value=1, max_value=200, value=10, key="sf_limit")

        sf_custom_soql = None
        if sf_qtype == "custom":
            sf_custom_soql = st.text_area(
                "SOQL query",
                placeholder="SELECT Id, Name FROM Contact WHERE ...",
                key="sf_custom",
            )

        if st.button("Run Query", key="sf_run", type="primary"):
            if not sf_secret:
                st.error("Enter a customer name first.")
            elif not sf_object and sf_qtype != "custom":
                st.error("Enter a Salesforce object name.")
            elif sf_qtype == "custom" and not sf_custom_soql:
                st.error("Enter a SOQL query.")
            else:
                try:
                    client = SalesforceClient(session)
                    with st.spinner("Fetching credentials from AWS..."):
                        client.load_secret(sf_secret, sf_oauth)

                    if sf_qtype == "shape":
                        with st.spinner(f"Describing {sf_object}..."):
                            fields = client.describe(sf_object)
                        st.success(f"{len(fields)} fields found")
                        st.dataframe(fields, use_container_width=True)
                        excel_download_button(make_excel(fields, "Object Shape"), f"{sf_object}_shape_{ts()}.xlsx")

                    else:
                        if sf_custom_soql:
                            soql = sf_custom_soql
                        elif sf_qtype == "count":
                            soql = f"SELECT COUNT() FROM {sf_object}"
                        elif sf_qtype == "list":
                            soql = f"SELECT Id, Name FROM {sf_object} LIMIT 20"
                        elif sf_qtype == "all":
                            soql = f"SELECT FIELDS(ALL) FROM {sf_object} LIMIT {sf_limit}"
                        else:
                            soql = f"SELECT Id FROM {sf_object} LIMIT 10"

                        with st.spinner("Running query..."):
                            result = client.query(soql)

                        records = result.get("records", [])
                        total = result.get("totalSize", 0)

                        if sf_qtype == "count":
                            count_val = records[0]["expr0"] if records and "expr0" in records[0] else total
                            st.metric("Total records", f"{count_val:,}")
                        else:
                            clean = [{k: v for k, v in r.items() if k != "attributes"} for r in records]
                            st.success(f"{total:,} total — {len(clean)} returned")
                            st.dataframe(clean, use_container_width=True)
                            if sf_qtype in ("all", "custom"):
                                excel_download_button(
                                    make_excel(clean, "Query Results"),
                                    f"{sf_object or 'query'}_results_{ts()}.xlsx",
                                )

                except Exception as exc:
                    st.error(f"Error: {exc}")
