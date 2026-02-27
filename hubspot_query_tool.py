#!/usr/bin/env python3
"""
HubSpot Query Tool
Fetches credentials from AWS Secrets Manager and queries HubSpot CRM objects

USAGE: Edit the CONFIG section below and run: python hubspot_query_tool.py
"""

import json
import boto3
import requests
import os
from typing import Dict, Any, List, Optional
from datetime import datetime
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ============================================================================
# CONFIGURATION - EDIT THIS SECTION
#   source .venv/bin/activate
#   aws sso login --profile hook-production-tic
#   python3 hubspot_query_tool.py
# ============================================================================

CONFIG = {
    # AWS Settings
    'aws_profile': 'hook-production-tic',
    'aws_region': 'eu-west-1',

    # Secret path in format: customer/integration
    # Example: 'conga/hubspot'
    'secret_path': 'opiniion/hubspot',

    # HubSpot object type to query
    # Standard: 'contacts', 'companies', 'deals', 'tickets',
    #           'line_items', 'products', 'quotes', 'calls', 'emails',
    #           'meetings', 'notes', 'tasks', 'communications'
    # Custom:   use the fully qualified name (e.g., 'p12345_my_object')
    'object_type': 'contacts',

    # Query type: 'count', 'list', 'all', 'shape', or 'search'
    # - 'count': Returns total record count for the object
    # - 'list':  Returns id + default properties (up to query_limit records)
    # - 'all':   Returns ALL properties for records -- saves as .xlsx
    # - 'shape': Returns all property names, types, and labels -- saves as .xlsx
    # - 'search': Filter records using search_filters below
    'query_type': 'all',

    # Number of records to fetch for 'list', 'all', and 'search' types
    # HubSpot max per page is 100; tool paginates automatically up to this total
    'query_limit': 500,

    # Object discovery mode: set to True to list all available HubSpot object types
    # Shows standard objects + custom schemas, with record counts for each
    'search_objects_mode': True,
    'search_objects_filter': 'com',  # Filter by name/label, or None to list all

    # Search filters (only used if query_type is 'search')
    # All filters in this list are ANDed together
    # Operators: EQ, NEQ, LT, LTE, GT, GTE, HAS_PROPERTY, NOT_HAS_PROPERTY,
    #            CONTAINS_TOKEN, NOT_CONTAINS_TOKEN, BETWEEN, IN, NOT_IN
    'search_filters': [
        # {'propertyName': 'lifecyclestage', 'operator': 'EQ', 'value': 'customer'},
        # {'propertyName': 'email', 'operator': 'CONTAINS_TOKEN', 'value': '@example.com'},
    ],

    # Specific properties to include (list of API property names)
    # Leave empty [] to use HubSpot defaults for 'list', or ALL for 'all'/'search'
    # Example: ['firstname', 'lastname', 'email', 'lifecyclestage']
    'properties': [],

    # Always refresh the access token via OAuth before running.
    # Recommended: True — HubSpot tokens expire after 30 minutes.
    # Set to False only if your secret contains a long-lived private app token.
    'always_refresh_token': True,

   
}

# ============================================================================
# END CONFIGURATION
# ============================================================================

STANDARD_OBJECTS = [
    'contacts', 'companies', 'deals', 'tickets',
    'line_items', 'products', 'quotes', 'calls',
    'emails', 'meetings', 'notes', 'tasks', 'communications',
]

BASE_URL = 'https://api.hubapi.com'

console = Console()


class HubSpotQueryTool:
    def __init__(self, profile: str, region: str):
        self.profile = profile
        self.region = region
        self.session = boto3.Session(profile_name=profile, region_name=region)
        self.secrets_client = self.session.client('secretsmanager')
        self.token = None
        self.auth_type = None  # 'bearer' or 'hapikey'

    def get_secret(self, secret_name: str) -> Dict[str, Any]:
        """Fetch secret from AWS Secrets Manager"""
        try:
            console.print(f"[cyan]Fetching secret: {secret_name}[/cyan]")
            response = self.secrets_client.get_secret_value(SecretId=secret_name)
            return json.loads(response['SecretString'])
        except Exception as e:
            console.print(f"[red]Error fetching secret: {str(e)}[/red]")
            raise

    def authenticate(self, credentials: Dict[str, Any], always_refresh: bool = True):
        """Determine auth method and set self.token / self.auth_type.

        Auth priority:
        1. hapikey  → legacy API key, passed as ?hapikey= query param
        2. OAuth refresh (if refresh_token + client_id + client_secret present and always_refresh=True)
        3. Stored access_token / token / api_key as Bearer
        """
        # Legacy API key — different auth mechanism, can't use Bearer
        if credentials.get('hapikey'):
            self.token = credentials['hapikey']
            self.auth_type = 'hapikey'
            console.print("[cyan]Auth: legacy API key (hapikey)[/cyan]")
            return

        # OAuth refresh — recommended when token expires in 30 min
        has_oauth = all(k in credentials for k in ('client_id', 'client_secret', 'refresh_token'))
        if always_refresh and has_oauth:
            console.print("[cyan]Refreshing OAuth access token...[/cyan]")
            self.token = self._refresh_oauth_token(credentials)
            self.auth_type = 'bearer'
            console.print("[green]✓ Fresh access token obtained[/green]")
            return

        # Fall back to stored token
        for key in ('access_token', 'token', 'api_key'):
            if credentials.get(key):
                self.token = credentials[key]
                self.auth_type = 'bearer'
                console.print(f"[cyan]Auth: stored Bearer token (key '{key}')[/cyan]")
                return

        raise ValueError(
            f"No token found in credentials. "
            f"Expected one of: access_token, hapikey, token, api_key, or OAuth fields "
            f"(client_id + client_secret + refresh_token). "
            f"Found keys: {list(credentials.keys())}"
        )

    def _refresh_oauth_token(self, credentials: Dict[str, Any]) -> str:
        """Get a fresh HubSpot access token using the refresh_token grant"""
        url = "https://api.hubapi.com/oauth/v1/token"
        data = {
            'grant_type': 'refresh_token',
            'client_id': credentials['client_id'],
            'client_secret': credentials['client_secret'],
            'refresh_token': credentials['refresh_token'],
        }
        try:
            response = requests.post(
                url, data=data,
                headers={'Content-Type': 'application/x-www-form-urlencoded'}
            )
            response.raise_for_status()
            return response.json()['access_token']
        except Exception as e:
            console.print(f"[red]OAuth token refresh failed: {str(e)}[/red]")
            if hasattr(e, 'response') and e.response is not None:
                console.print(f"[red]Response: {e.response.text}[/red]")
            raise

    def _get(self, url: str, extra_params: Dict = None) -> requests.Response:
        """Authenticated GET request"""
        headers = {'Content-Type': 'application/json'}
        params = {}
        if self.auth_type == 'bearer':
            headers['Authorization'] = f'Bearer {self.token}'
        else:  # hapikey
            params['hapikey'] = self.token
        if extra_params:
            params.update(extra_params)
        return requests.get(url, headers=headers, params=params)

    def _post(self, url: str, payload: Dict, extra_params: Dict = None) -> requests.Response:
        """Authenticated POST request"""
        headers = {'Content-Type': 'application/json'}
        params = {}
        if self.auth_type == 'bearer':
            headers['Authorization'] = f'Bearer {self.token}'
        else:  # hapikey
            params['hapikey'] = self.token
        if extra_params:
            params.update(extra_params)
        return requests.post(url, headers=headers, params=params, json=payload)

    def get_all_properties(self, object_type: str) -> List[Dict]:
        """Fetch all properties for a given object type"""
        url = f"{BASE_URL}/crm/v3/properties/{object_type}"
        try:
            console.print(f"[cyan]Fetching properties for: {object_type}[/cyan]")
            response = self._get(url)
            response.raise_for_status()
            return response.json().get('results', [])
        except Exception as e:
            console.print(f"[red]Error fetching properties: {str(e)}[/red]")
            if hasattr(e, 'response') and e.response is not None:
                console.print(f"[red]Response: {e.response.text}[/red]")
            raise

    def count_records(self, object_type: str) -> int:
        """Get total record count for an object type"""
        url = f"{BASE_URL}/crm/v3/objects/{object_type}/search"
        payload = {'filterGroups': [], 'limit': 1, 'properties': ['hs_object_id']}
        try:
            response = self._post(url, payload)
            response.raise_for_status()
            return response.json().get('total', 0)
        except Exception:
            return -1

    def list_records(self, object_type: str, properties: List[str] = None, limit: int = 20) -> List[Dict]:
        """List records using the basic GET endpoint"""
        url = f"{BASE_URL}/crm/v3/objects/{object_type}"
        extra_params = {'limit': min(limit, 100)}
        if properties:
            extra_params['properties'] = ','.join(properties)
        try:
            console.print(f"[cyan]Fetching {object_type} records...[/cyan]")
            response = self._get(url, extra_params)
            response.raise_for_status()
            return response.json().get('results', [])
        except Exception as e:
            console.print(f"[red]Error listing records: {str(e)}[/red]")
            if hasattr(e, 'response') and e.response is not None:
                console.print(f"[red]Response: {e.response.text}[/red]")
            raise

    def search_records(self, object_type: str, filters: List[Dict],
                       properties: List[str] = None, limit: int = 100) -> Dict[str, Any]:
        """Search records using the HubSpot search endpoint with filters"""
        url = f"{BASE_URL}/crm/v3/objects/{object_type}/search"
        payload = {
            'filterGroups': [{'filters': filters}] if filters else [],
            'limit': min(limit, 100),
        }
        if properties:
            payload['properties'] = properties
        try:
            console.print(f"[cyan]Searching {object_type} records...[/cyan]")
            response = self._post(url, payload)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            console.print(f"[red]Error searching records: {str(e)}[/red]")
            if hasattr(e, 'response') and e.response is not None:
                console.print(f"[red]Response: {e.response.text}[/red]")
            raise

    def fetch_all_records(self, object_type: str, all_property_names: List[str], limit: int) -> List[Dict]:
        """Fetch records with all properties, paginating up to limit"""
        url = f"{BASE_URL}/crm/v3/objects/{object_type}/search"
        records = []
        after = None

        while len(records) < limit:
            batch_size = min(100, limit - len(records))
            payload = {
                'filterGroups': [],
                'properties': all_property_names,
                'limit': batch_size,
            }
            if after:
                payload['after'] = after

            response = self._post(url, payload)
            response.raise_for_status()
            data = response.json()

            batch = data.get('results', [])
            records.extend(batch)

            paging = data.get('paging', {})
            after = paging.get('next', {}).get('after')
            if not after or not batch:
                break

        return records

    def get_object_schemas(self) -> List[Dict]:
        """Get all custom object schemas"""
        url = f"{BASE_URL}/crm/v3/schemas"
        try:
            response = self._get(url)
            response.raise_for_status()
            return response.json().get('results', [])
        except Exception as e:
            console.print(f"[yellow]Could not fetch custom schemas: {str(e)}[/yellow]")
            return []

    def flatten_record(self, record: Dict) -> Dict:
        """Flatten HubSpot record structure: {id, properties: {...}} → one flat dict"""
        flat = {'id': record.get('id', '')}
        flat.update(record.get('properties', {}))
        return flat

    def save_to_excel(self, records: List[Dict], object_type: str, save_directory: str) -> Optional[str]:
        """Save query results to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Query Results"

        if not records:
            console.print("[yellow]No records to save[/yellow]")
            return None

        flat_records = [self.flatten_record(r) for r in records]

        all_fields = set()
        for record in flat_records:
            all_fields.update(record.keys())
        all_fields = ['id'] + sorted(f for f in all_fields if f != 'id')

        # HubSpot orange header
        header_fill = PatternFill(start_color="FF7A00", end_color="FF7A00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, field in enumerate(all_fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, record in enumerate(flat_records, 2):
            for col_idx, field in enumerate(all_fields, 1):
                value = record.get(field, '')
                if isinstance(value, dict):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{object_type}_records_{timestamp}.xlsx"
        filepath = os.path.join(save_directory, filename)
        wb.save(filepath)
        return filepath

    def save_shape_to_excel(self, properties: List[Dict], object_type: str, save_directory: str) -> Optional[str]:
        """Save property schema to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Object Shape"

        if not properties:
            console.print("[yellow]No properties found[/yellow]")
            return None

        header_fill = PatternFill(start_color="FF7A00", end_color="FF7A00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ['Property Name', 'Label', 'Type', 'Field Type', 'Group']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, prop in enumerate(properties, 2):
            ws.cell(row=row_idx, column=1, value=prop.get('name', ''))
            ws.cell(row=row_idx, column=2, value=prop.get('label', ''))
            ws.cell(row=row_idx, column=3, value=prop.get('type', ''))
            ws.cell(row=row_idx, column=4, value=prop.get('fieldType', ''))
            ws.cell(row=row_idx, column=5, value=prop.get('groupName', ''))

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{object_type}_shape_{timestamp}.xlsx"
        filepath = os.path.join(save_directory, filename)
        wb.save(filepath)
        return filepath

    def display_records(self, records: List[Dict], title: str = "Results"):
        """Display records in a rich table (max 20 rows)"""
        if not records:
            console.print("[yellow]No records returned[/yellow]")
            return

        flat_records = [self.flatten_record(r) for r in records]

        all_fields = set()
        for r in flat_records:
            all_fields.update(r.keys())
        sorted_fields = ['id'] + sorted(f for f in all_fields if f != 'id')

        table = Table(title=title, show_lines=True)
        for field in sorted_fields:
            table.add_column(field, style="cyan")

        display_count = min(20, len(flat_records))
        for record in flat_records[:display_count]:
            row = [str(record.get(field, '')) for field in sorted_fields]
            table.add_row(*row)

        console.print(table)

        if len(flat_records) > display_count:
            console.print(f"\n[yellow]Showing {display_count} of {len(flat_records)} records[/yellow]")

    def run_query(self, secret_path: str, object_type: str, query_type: str,
                  search_filters: List[Dict], properties: List[str],
                  query_limit: int, always_refresh: bool = True, save_directory: str = None):
        """Main execution flow"""
        try:
            # Ask for save directory if needed
            if query_type.lower() in ('all', 'shape') and not save_directory:
                console.print("[yellow]Query type 'all' or 'shape' will export results to Excel.[/yellow]")
                save_directory = input("Enter directory to save Excel file (or press Enter for current): ").strip()
                if not save_directory:
                    save_directory = os.getcwd()
                if not os.path.isdir(save_directory):
                    console.print(f"[red]Error: Directory does not exist: {save_directory}[/red]")
                    return
                console.print(f"[green]Will save to: {save_directory}[/green]\n")

            # Display config
            config_table = Table(title="Current Configuration", show_header=False)
            config_table.add_column("Setting", style="cyan")
            config_table.add_column("Value", style="green")
            config_table.add_row("AWS Profile", self.profile)
            config_table.add_row("AWS Region", self.region)
            config_table.add_row("Secret Path", secret_path)
            config_table.add_row("Object Type", object_type)
            config_table.add_row("Query Type", query_type)
            config_table.add_row("Limit", str(query_limit))
            config_table.add_row("Always Refresh Token", "Enabled" if always_refresh else "Disabled")
            if query_type.lower() in ('all', 'shape') and save_directory:
                config_table.add_row("Save Directory", save_directory)
            console.print(config_table)
            console.print()

            # Fetch credentials and authenticate
            with console.status("[bold green]Fetching credentials from AWS..."):
                credentials = self.get_secret(secret_path)
            self.authenticate(credentials, always_refresh=always_refresh)
            console.print()

            # --- SHAPE ---
            if query_type.lower() == 'shape':
                with console.status(f"[bold green]Fetching properties for {object_type}..."):
                    props = self.get_all_properties(object_type)

                table = Table(title=f"{object_type} - Properties ({len(props)})", show_lines=True)
                table.add_column("#", style="dim", justify="right")
                table.add_column("Property Name", style="cyan")
                table.add_column("Label", style="green")
                table.add_column("Type", style="yellow")
                table.add_column("Field Type", style="blue")
                table.add_column("Group", style="magenta")

                for idx, prop in enumerate(props, 1):
                    table.add_row(
                        str(idx),
                        prop.get('name', ''),
                        prop.get('label', ''),
                        prop.get('type', ''),
                        prop.get('fieldType', ''),
                        prop.get('groupName', ''),
                    )
                console.print(table)

                filepath = self.save_shape_to_excel(props, object_type, save_directory)
                if filepath:
                    console.print(f"\n[bold green]✓ Shape saved to Excel![/bold green]")
                    console.print(f"[cyan]File:[/cyan] {filepath}")
                    console.print(f"[cyan]Total properties:[/cyan] {len(props)}\n")
                return

            # --- COUNT ---
            if query_type.lower() == 'count':
                with console.status(f"[bold green]Counting {object_type} records..."):
                    total = self.count_records(object_type)

                panel = Panel(
                    f"[green]Total Records:[/green] [bold]{total}[/bold]\n"
                    f"[cyan]Object:[/cyan] {object_type}",
                    title="[bold]Count Result[/bold]",
                    border_style="blue"
                )
                console.print(panel)
                return

            # --- LIST ---
            if query_type.lower() == 'list':
                props_to_fetch = properties if properties else None
                with console.status(f"[bold green]Fetching {object_type} records..."):
                    records = self.list_records(object_type, props_to_fetch, query_limit)

                panel = Panel(
                    f"[green]Records Returned:[/green] {len(records)}\n"
                    f"[cyan]Object:[/cyan] {object_type}",
                    title="[bold]List Results[/bold]",
                    border_style="blue"
                )
                console.print(panel)
                self.display_records(records, title=f"{object_type} Records")
                return

            # --- ALL ---
            if query_type.lower() == 'all':
                with console.status(f"[bold green]Fetching all properties for {object_type}..."):
                    all_props = self.get_all_properties(object_type)
                all_prop_names = [p['name'] for p in all_props]
                console.print(f"[cyan]Found {len(all_prop_names)} properties, fetching records...[/cyan]")

                with console.status(f"[bold green]Fetching {object_type} records..."):
                    records = self.fetch_all_records(object_type, all_prop_names, query_limit)

                console.print(f"[cyan]Saving {len(records)} records to Excel...[/cyan]")
                filepath = self.save_to_excel(records, object_type, save_directory)
                if filepath:
                    console.print(f"\n[bold green]✓ Results saved![/bold green]")
                    console.print(f"[cyan]File:[/cyan] {filepath}")
                    console.print(f"[cyan]Records:[/cyan] {len(records)}")
                    console.print(f"[cyan]Properties:[/cyan] {len(all_prop_names)}\n")
                return

            # --- SEARCH ---
            if query_type.lower() == 'search':
                props_to_fetch = properties if properties else None
                with console.status(f"[bold green]Searching {object_type} records..."):
                    result = self.search_records(object_type, search_filters, props_to_fetch, query_limit)
                records = result.get('results', [])
                total = result.get('total', len(records))

                panel = Panel(
                    f"[green]Total Matching:[/green] {total}\n"
                    f"[green]Records Returned:[/green] {len(records)}\n"
                    f"[cyan]Object:[/cyan] {object_type}\n"
                    f"[cyan]Filters:[/cyan] {search_filters}",
                    title="[bold]Search Results[/bold]",
                    border_style="blue"
                )
                console.print(panel)
                self.display_records(records, title=f"{object_type} Search Results")
                return

        except Exception as e:
            console.print(f"[bold red]Error: {str(e)}[/bold red]")
            raise

    def run_search_objects(self, secret_path: str, search_filter: str = None, always_refresh: bool = True):
        """List all available HubSpot object types with record counts"""
        try:
            with console.status("[bold green]Fetching credentials from AWS..."):
                credentials = self.get_secret(secret_path)
            self.authenticate(credentials, always_refresh=always_refresh)
            console.print()

            # Combine standard objects + custom schemas
            with console.status("[bold green]Fetching custom object schemas..."):
                custom_schemas = self.get_object_schemas()

            all_objects = [{'name': o, 'label': o.title(), 'type': 'standard'} for o in STANDARD_OBJECTS]
            for schema in custom_schemas:
                all_objects.append({
                    'name': schema.get('fullyQualifiedName', schema.get('name', '')),
                    'label': schema.get('labels', {}).get('singular', schema.get('name', '')),
                    'type': 'custom',
                })

            # Apply filter
            if search_filter:
                term = search_filter.lower()
                filtered = [o for o in all_objects if term in o['name'].lower() or term in o['label'].lower()]
                console.print(f"[cyan]Found {len(filtered)} object(s) matching '{search_filter}'[/cyan]\n")
            else:
                filtered = all_objects
                console.print(f"[cyan]{len(filtered)} object type(s) available[/cyan]\n")

            if not filtered:
                console.print(f"[yellow]No objects found matching: {search_filter}[/yellow]")
                return

            table = Table(title="HubSpot Object Types", show_lines=True)
            table.add_column("Object Name", style="cyan", no_wrap=True)
            table.add_column("Label", style="green")
            table.add_column("Type", style="magenta")
            table.add_column("Record Count", style="yellow", justify="right")

            with console.status("[bold green]Counting records...") as status:
                for idx, obj in enumerate(filtered, 1):
                    status.update(f"[bold green]Counting records... ({idx}/{len(filtered)}) {obj['name']}")
                    count = self.count_records(obj['name'])
                    count_str = str(count) if count >= 0 else "Error"
                    table.add_row(obj['name'], obj['label'], obj['type'], count_str)

            console.print(table)
            console.print(f"\n[green]Displayed {len(filtered)} object(s)[/green]\n")

        except Exception as e:
            console.print(f"[bold red]Error: {str(e)}[/bold red]")
            raise


def main():
    console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]         HubSpot Query Tool via AWS Secrets           [/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

    tool = HubSpotQueryTool(
        profile=CONFIG['aws_profile'],
        region=CONFIG['aws_region'],
    )

    if CONFIG['search_objects_mode']:
        if not CONFIG['secret_path']:
            console.print("[red]Error: 'secret_path' is not set in CONFIG[/red]")
            console.print("[yellow]Please edit the CONFIG section at the top of this file[/yellow]")
            return
        tool.run_search_objects(
            secret_path=CONFIG['secret_path'],
            search_filter=CONFIG['search_objects_filter'],
            always_refresh=CONFIG['always_refresh_token'],
        )
        return

    if not CONFIG['secret_path']:
        console.print("[red]Error: 'secret_path' is not set in CONFIG[/red]")
        console.print("[yellow]Please edit the CONFIG section at the top of this file[/yellow]")
        return

    if not CONFIG['object_type']:
        console.print("[red]Error: 'object_type' is not set in CONFIG[/red]")
        console.print("[yellow]Please edit the CONFIG section at the top of this file[/yellow]")
        return

    tool.run_query(
        secret_path=CONFIG['secret_path'],
        object_type=CONFIG['object_type'],
        query_type=CONFIG['query_type'],
        search_filters=CONFIG['search_filters'],
        properties=CONFIG['properties'],
        query_limit=CONFIG['query_limit'],
        always_refresh=CONFIG['always_refresh_token'],
    )


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        console.print("\n[yellow]Interrupted by user[/yellow]")
    except Exception as e:
        console.print(f"\n[bold red]Fatal error: {str(e)}[/bold red]")
