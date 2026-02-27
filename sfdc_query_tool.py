#!/usr/bin/env python3
"""
Salesforce Query Tool
Fetches credentials from AWS Secrets Manager and queries Salesforce objects

USAGE: Edit the CONFIG section below and run: python salesforce_query_tool.py
"""

import json
import boto3
import requests
import os
from typing import Dict, Any
from datetime import datetime
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich import print as rprint
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ============================================================================
# CONFIGURATION - EDIT THIS SECTION 
#   source .venv/bin/activate
#   aws sso login --profile hook-production-tic
#   python3 sfdc_query_tool.py  
# ============================================================================

CONFIG = {
    # AWS Settings
    'aws_profile': 'hook-production-tic',
    'aws_region': 'eu-west-1',

    # Secret path in format: customer/integration
    # Example: 'conga/salesforce'
    'secret_path': 'conga/salesforce',

    # Salesforce object to query
    # Examples: 'Asset', 'Account', 'Contact', 'Opportunity'
    'sobject': 'Apttus_Config2__AssetLineItem__Share',

    # Query type: 'count', 'list', 'all', 'shape', or 'custom'
    # - 'count': Returns count of records
    # - 'list': Returns Id and Name fields (max 20 records)
    # - 'all': Returns all fields (max 10 records) -- saves as .xlsx
    # - 'shape': Returns all column names and their data types for the object (no record data) -- saves as .xlsx
    # - 'custom': Use the custom_query below
    'query_type': 'all',

    # Limit for 'all' query type (if None or empty, defaults to 10)
    # MAX 200 !!!
    'all_query_limit': 50,

    # Custom SOQL query (only used if query_type is 'custom')
    # Example: "SELECT Id, Name, Status FROM Asset WHERE Status = 'Active'"
    'custom_query': None,

    # Object discovery mode: set to True to list and search Salesforce objects
    # When True, set 'search_objects_filter' to search term (or None to list all)
    'search_objects_mode': True,
    'search_objects_filter': 'opportuni',  # Example: 'asset', 'account', 'contact'

    # Auto-refresh credentials: If True, automatically refetch from AWS when token expires
    # This assumes your AWS secret is kept up-to-date with fresh credentials
    'auto_refresh_on_expire': True,

    # Always use OAuth flow: If True, always generate fresh token via OAuth instead of using stored token
    # Recommended if your stored access_token in AWS Secrets Manager is often stale
    'always_use_oauth': True,

}

# ============================================================================
# END CONFIGURATION
# ============================================================================


console = Console()


class SalesforceQueryTool:
    def __init__(self, profile: str, region: str):
        self.profile = profile
        self.region = region
        self.session = boto3.Session(profile_name=profile, region_name=region)
        self.secrets_client = self.session.client('secretsmanager')

    def get_secret(self, secret_name: str, force_refresh: bool = False) -> Dict[str, Any]:
        """Fetch secret from AWS Secrets Manager"""
        try:
            console.print(f"[cyan]Fetching secret: {secret_name}[/cyan]")
            response = self.secrets_client.get_secret_value(SecretId=secret_name)
            secret_string = response['SecretString']
            credentials = json.loads(secret_string)

            # If force_refresh is True, fetch fresh credentials again
            if force_refresh:
                console.print(f"[yellow]Refreshing credentials from AWS...[/yellow]")
                response = self.secrets_client.get_secret_value(SecretId=secret_name)
                secret_string = response['SecretString']
                credentials = json.loads(secret_string)

            return credentials
        except Exception as e:
            console.print(f"[red]Error fetching secret: {str(e)}[/red]")
            raise

    def get_salesforce_objects(self, instance_url: str, access_token: str) -> list:
        """Get list of all Salesforce objects"""
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }

        instance_url = instance_url.rstrip('/')
        sobjects_url = f"{instance_url}/services/data/v59.0/sobjects"

        try:
            console.print("[cyan]Fetching all Salesforce objects...[/cyan]")
            response = requests.get(sobjects_url, headers=headers)
            response.raise_for_status()
            return response.json().get('sobjects', [])
        except Exception as e:
            console.print(f"[red]Error fetching Salesforce objects: {str(e)}[/red]")
            if hasattr(e, 'response') and e.response is not None:
                console.print(f"[red]Response: {e.response.text}[/red]")
            raise

    def describe_sobject(self, instance_url: str, access_token: str, sobject: str) -> list:
        """Describe a Salesforce object to get all field names and data types"""
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }

        instance_url = instance_url.rstrip('/')
        describe_url = f"{instance_url}/services/data/v59.0/sobjects/{sobject}/describe"

        try:
            console.print(f"[cyan]Describing object: {sobject}[/cyan]")
            response = requests.get(describe_url, headers=headers)
            response.raise_for_status()
            data = response.json()
            fields = data.get('fields', [])
            return [{'name': f['name'], 'type': f['type'], 'label': f['label'], 'length': f['length']} for f in fields]
        except Exception as e:
            console.print(f"[red]Error describing Salesforce object: {str(e)}[/red]")
            if hasattr(e, 'response') and e.response is not None:
                console.print(f"[red]Response: {e.response.text}[/red]")
            raise

    def count_records(self, instance_url: str, access_token: str, sobject: str) -> int:
        """Count records in a Salesforce object"""
        try:
            query = f"SELECT COUNT() FROM {sobject}"
            result = self.query_salesforce(instance_url, access_token, query, silent=True)
            return result.get('totalSize', 0)
        except Exception:
            # Silently return -1 for errors during bulk counting
            return -1  # Return -1 to indicate error

    def get_salesforce_access_token(self, credentials: Dict[str, Any], force_refresh: bool = False) -> str:
        """Get access token using OAuth credentials"""
        # Check if access_token is already in the secret and we're not forcing refresh
        if 'access_token' in credentials and not force_refresh:
            return credentials['access_token']

        instance_url = credentials.get('instance_url', 'https://login.salesforce.com')
        token_url = f"{instance_url.rstrip('/')}/services/oauth2/token"

        # Try refresh token flow first if available
        if 'refresh_token' in credentials:
            console.print("[cyan]Attempting to refresh access token...[/cyan]")
            data = {
                'grant_type': 'refresh_token',
                'client_id': credentials['client_id'],
                'client_secret': credentials['client_secret'],
                'refresh_token': credentials['refresh_token']
            }

            try:
                response = requests.post(token_url, data=data)
                response.raise_for_status()
                new_token = response.json()['access_token']
                console.print("[green]✓ Successfully refreshed access token[/green]")
                return new_token
            except Exception as e:
                console.print(f"[yellow]Refresh token flow failed: {str(e)}[/yellow]")

        # Try password flow if username and password are available
        if 'username' in credentials and 'password' in credentials:
            console.print("[cyan]Attempting password authentication...[/cyan]")
            data = {
                'grant_type': 'password',
                'client_id': credentials['client_id'],
                'client_secret': credentials['client_secret'],
                'username': credentials['username'],
                'password': credentials['password']
            }

            # Add security token if present
            if 'security_token' in credentials:
                data['password'] = credentials['password'] + credentials['security_token']

            try:
                response = requests.post(token_url, data=data)
                response.raise_for_status()
                new_token = response.json()['access_token']
                console.print("[green]✓ Successfully authenticated with password[/green]")
                return new_token
            except Exception as e:
                console.print(f"[yellow]Password authentication failed: {str(e)}[/yellow]")

        # Try client credentials flow as last resort
        console.print("[cyan]Attempting client credentials flow...[/cyan]")
        data = {
            'grant_type': 'client_credentials',
            'client_id': credentials['client_id'],
            'client_secret': credentials['client_secret']
        }

        try:
            response = requests.post(token_url, data=data)
            response.raise_for_status()
            return response.json()['access_token']
        except Exception as e:
            console.print(f"[red]Error getting access token: {str(e)}[/red]")
            if hasattr(e, 'response') and e.response is not None:
                console.print(f"[red]Response: {e.response.text}[/red]")
            raise

    def query_salesforce(self, instance_url: str, access_token: str, soql_query: str, silent: bool = False) -> Dict[str, Any]:
        """Execute SOQL query against Salesforce"""
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }

        # Ensure instance_url doesn't have trailing slash
        instance_url = instance_url.rstrip('/')

        query_url = f"{instance_url}/services/data/v59.0/query"

        try:
            if not silent:
                console.print(f"[cyan]Executing query: {soql_query}[/cyan]")
            response = requests.get(query_url, headers=headers, params={'q': soql_query})
            response.raise_for_status()
            return response.json()
        except Exception as e:
            if not silent:
                console.print(f"[red]Error querying Salesforce: {str(e)}[/red]")
                if hasattr(e, 'response') and e.response is not None:
                    console.print(f"[red]Response: {e.response.text}[/red]")
            raise

    def save_to_excel(self, results: Dict[str, Any], query: str, save_directory: str, sobject: str) -> str:
        """Save query results to Excel file"""
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Query Results"

        # Get records
        records = results.get('records', [])
        if not records:
            console.print("[yellow]No records to save[/yellow]")
            return None

        # Get all unique fields (excluding attributes)
        all_fields = set()
        for record in records:
            all_fields.update(k for k in record.keys() if k != 'attributes')
        all_fields = sorted(all_fields)

        # Write header row with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, field in enumerate(all_fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.fill = header_fill
            cell.font = header_font

        # Write data rows
        for row_idx, record in enumerate(records, 2):
            for col_idx, field in enumerate(all_fields, 1):
                value = record.get(field, '')
                # Handle nested objects/dicts
                if isinstance(value, dict):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{sobject}_all_fields_{timestamp}.xlsx"
        filepath = os.path.join(save_directory, filename)

        # Save file
        wb.save(filepath)
        return filepath

    def save_shape_to_excel(self, fields: list, save_directory: str, sobject: str) -> str:
        """Save object shape (field names and types) to Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Object Shape"

        if not fields:
            console.print("[yellow]No fields found[/yellow]")
            return None

        # Write header row with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ['Field Name', 'Data Type', 'Label', 'Length']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Write data rows
        for row_idx, field in enumerate(fields, 2):
            ws.cell(row=row_idx, column=1, value=field['name'])
            ws.cell(row=row_idx, column=2, value=field['type'])
            ws.cell(row=row_idx, column=3, value=field['label'])
            ws.cell(row=row_idx, column=4, value=field['length'])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{sobject}_shape_{timestamp}.xlsx"
        filepath = os.path.join(save_directory, filename)

        wb.save(filepath)
        return filepath

    def display_results(self, results: Dict[str, Any], query: str):
        """Display query results in a nice format"""
        console.print("\n")

        # Summary Panel
        summary = Panel(
            f"[green]Total Records:[/green] {results.get('totalSize', 0)}\n"
            f"[cyan]Query:[/cyan] {query}",
            title="[bold]Query Summary[/bold]",
            border_style="blue"
        )
        console.print(summary)

        # Records Table
        if results.get('records'):
            records = results['records']

            # If it's a COUNT query, display differently
            if 'expr0' in records[0]:
                console.print(f"\n[bold green]Count Result: {records[0]['expr0']}[/bold green]")
            else:
                # Create table with fields from first record
                table = Table(title="Query Results", show_lines=True)

                # Get all unique fields (excluding attributes)
                all_fields = set()
                for record in records:
                    all_fields.update(k for k in record.keys() if k != 'attributes')

                # Add columns
                for field in sorted(all_fields):
                    table.add_column(field, style="cyan")

                # Add rows (limit to first 20 for readability)
                display_count = min(20, len(records))
                for record in records[:display_count]:
                    row = [str(record.get(field, '')) for field in sorted(all_fields)]
                    table.add_row(*row)

                console.print(table)

                if len(records) > display_count:
                    console.print(f"\n[yellow]Showing {display_count} of {len(records)} records[/yellow]")

        console.print("\n")

    def run_query(self, secret_path: str, sobject: str, query_type: str, custom_query: str = None,
                  auto_refresh: bool = True, always_use_oauth: bool = False, save_directory: str = None):
        """Main execution flow for querying"""
        try:
            # If query type is 'all' or 'shape' and no save directory provided, ask for it
            if query_type.lower() in ('all', 'shape') and not save_directory:
                console.print("[yellow]Query type 'all' or 'shape' will export results to Excel file.[/yellow]")
                save_directory = input("Enter directory path to save Excel file (or press Enter for current directory): ").strip()
                if not save_directory:
                    save_directory = os.getcwd()

                # Validate directory exists
                if not os.path.isdir(save_directory):
                    console.print(f"[red]Error: Directory does not exist: {save_directory}[/red]")
                    return

                console.print(f"[green]Will save to: {save_directory}[/green]\n")

            # Display configuration
            config_table = Table(title="Current Configuration", show_header=False)
            config_table.add_column("Setting", style="cyan")
            config_table.add_column("Value", style="green")
            config_table.add_row("AWS Profile", self.profile)
            config_table.add_row("AWS Region", self.region)
            config_table.add_row("Secret Path", secret_path)
            config_table.add_row("Salesforce Object", sobject)
            config_table.add_row("Query Type", query_type)
            config_table.add_row("Auto-refresh", "Enabled" if auto_refresh else "Disabled")
            config_table.add_row("Always Use OAuth", "Enabled" if always_use_oauth else "Disabled")
            if query_type.lower() in ('all', 'shape'):
                config_table.add_row("Save Directory", save_directory)
            console.print(config_table)
            console.print("\n")

            # Fetch credentials
            with console.status("[bold green]Fetching credentials from AWS..."):
                credentials = self.get_secret(secret_path)

            # Display credentials summary (without showing sensitive data)
            cred_table = Table(title="Credentials Found", show_header=False)
            cred_table.add_column("Key", style="cyan")
            cred_table.add_column("Status", style="green")

            for key in ['client_id', 'client_secret', 'access_token', 'instance_url']:
                if key in credentials:
                    value = "✓ Present" if credentials[key] else "✗ Missing"
                    cred_table.add_row(key, value)

            console.print(cred_table)
            console.print("\n")

            # Get instance URL
            instance_url = credentials.get('instance_url')
            if not instance_url:
                raise ValueError("instance_url not found in credentials")

            # Get access token
            if always_use_oauth:
                console.print("[cyan]Generating fresh access token via OAuth flow...[/cyan]")
                access_token = self.get_salesforce_access_token(credentials, force_refresh=True)
            else:
                access_token = credentials.get('access_token')
                if not access_token:
                    console.print("[yellow]No access_token in secret, attempting OAuth flow...[/yellow]")
                    access_token = self.get_salesforce_access_token(credentials)

            # Handle 'shape' query type separately (uses describe API, not SOQL)
            if query_type.lower() == 'shape':
                with console.status(f"[bold green]Describing {sobject}..."):
                    fields = self.describe_sobject(instance_url, access_token, sobject)

                # Display in terminal
                table = Table(title=f"{sobject} - Object Shape ({len(fields)} fields)", show_lines=True)
                table.add_column("#", style="dim", justify="right")
                table.add_column("Field Name", style="cyan")
                table.add_column("Data Type", style="green")
                table.add_column("Label", style="yellow")
                table.add_column("Length", style="blue", justify="right")

                for idx, field in enumerate(fields, 1):
                    table.add_row(str(idx), field['name'], field['type'], field['label'], str(field['length']))

                console.print(table)

                # Save to Excel
                filepath = self.save_shape_to_excel(fields, save_directory, sobject)
                if filepath:
                    console.print(f"\n[bold green]✓ Shape saved to Excel![/bold green]")
                    console.print(f"[cyan]File location:[/cyan] {filepath}")
                    console.print(f"[cyan]Total fields:[/cyan] {len(fields)}\n")
                return

            # Build SOQL query
            if custom_query:
                soql_query = custom_query
            elif query_type.lower() == "count":
                soql_query = f"SELECT COUNT() FROM {sobject}"
            elif query_type.lower() == "list":
                soql_query = f"SELECT Id, Name FROM {sobject} LIMIT 20"
            elif query_type.lower() == "all":
                limit = CONFIG.get('all_query_limit') or 10
                soql_query = f"SELECT FIELDS(ALL) FROM {sobject} LIMIT {limit}"
            else:
                soql_query = f"SELECT Id FROM {sobject} LIMIT 10"

            # Execute query (with retry on auth failure if auto_refresh is enabled)
            try:
                results = self.query_salesforce(instance_url, access_token, soql_query)
            except requests.exceptions.HTTPError as e:
                # If we get 401 Unauthorized and auto_refresh is enabled, refetch the secret
                if e.response.status_code == 401 and auto_refresh:
                    console.print("[yellow]Access token expired, attempting to generate fresh token...[/yellow]")
                    try:
                        # Refetch the secret to get latest OAuth credentials
                        credentials = self.get_secret(secret_path, force_refresh=True)

                        # Always attempt OAuth flow to get a fresh token (don't trust stored access_token)
                        console.print("[cyan]Generating new access token using OAuth flow...[/cyan]")
                        access_token = self.get_salesforce_access_token(credentials, force_refresh=True)

                        # Update instance_url in case it changed
                        instance_url = credentials.get('instance_url')
                        if not instance_url:
                            raise ValueError("instance_url not found in refreshed credentials")

                        console.print("[cyan]Retrying query with fresh access token...[/cyan]")
                        results = self.query_salesforce(instance_url, access_token, soql_query)

                    except Exception as refresh_error:
                        console.print(f"[red]Failed to refresh credentials: {str(refresh_error)}[/red]")
                        console.print(
                            "\n[bold yellow]The credentials in AWS Secrets Manager may be stale.[/bold yellow]")
                        console.print(
                            "[bold yellow]Your shell script can generate fresh credentials. Consider running:[/bold yellow]")
                        console.print(f"   [cyan]aws_get_secret {secret_path}[/cyan]")
                        console.print("\n[bold yellow]Then update the secret in AWS if needed.[/bold yellow]")
                        raise
                else:
                    raise

            # Display or save results
            if query_type.lower() == 'all':
                # Save to Excel for 'all' query type
                console.print("[cyan]Saving results to Excel...[/cyan]")
                filepath = self.save_to_excel(results, soql_query, save_directory, sobject)
                if filepath:
                    console.print(f"\n[bold green]✓ Results saved successfully![/bold green]")
                    console.print(f"[cyan]File location:[/cyan] {filepath}")
                    console.print(f"[cyan]Total records:[/cyan] {results.get('totalSize', 0)}")

                    # Count columns
                    records = results.get('records', [])
                    if records:
                        all_fields = set()
                        for record in records:
                            all_fields.update(k for k in record.keys() if k != 'attributes')
                        console.print(f"[cyan]Total columns:[/cyan] {len(all_fields)}\n")
            else:
                # Display in terminal for other query types
                self.display_results(results, soql_query)

        except Exception as e:
            console.print(f"[bold red]Error: {str(e)}[/bold red]")
            raise

    def run_search_objects(self, secret_path: str, search_filter: str = None, always_use_oauth: bool = False):
        """Search Salesforce objects and display with record counts"""
        try:
            # Fetch credentials
            with console.status("[bold green]Fetching credentials from AWS..."):
                credentials = self.get_secret(secret_path)

            # Get instance URL
            instance_url = credentials.get('instance_url')
            if not instance_url:
                raise ValueError("instance_url not found in credentials")

            # Get access token
            if always_use_oauth:
                console.print("[cyan]Generating fresh access token via OAuth flow...[/cyan]")
                access_token = self.get_salesforce_access_token(credentials, force_refresh=True)
            else:
                access_token = credentials.get('access_token')
                if not access_token:
                    console.print("[yellow]No access_token in secret, attempting OAuth flow...[/yellow]")
                    access_token = self.get_salesforce_access_token(credentials)

            # Get all objects
            all_objects = self.get_salesforce_objects(instance_url, access_token)

            # Filter objects if search term provided
            if search_filter:
                search_term = search_filter.lower()
                filtered_objects = [
                    obj for obj in all_objects
                    if search_term in obj['name'].lower() or search_term in obj.get('label', '').lower()
                ]
                console.print(f"\n[cyan]Found {len(filtered_objects)} object(s) matching '{search_filter}'[/cyan]\n")
            else:
                filtered_objects = all_objects
                console.print(f"\n[cyan]Found {len(filtered_objects)} total object(s)[/cyan]\n")

            if not filtered_objects:
                console.print(f"[yellow]No objects found matching: {search_filter}[/yellow]")
                return

            # Create results table
            table = Table(title="Salesforce Objects", show_lines=True)
            table.add_column("Object Name", style="cyan", no_wrap=True)
            table.add_column("Label", style="green")
            table.add_column("Record Count", style="yellow", justify="right")
            table.add_column("Queryable", style="blue", justify="center")

            # Get counts for each object (with progress indicator)
            with console.status("[bold green]Counting records...") as status:
                for idx, obj in enumerate(filtered_objects, 1):
                    obj_name = obj['name']
                    obj_label = obj.get('label', '')
                    is_queryable = obj.get('queryable', False)

                    status.update(f"[bold green]Counting records... ({idx}/{len(filtered_objects)}) {obj_name}")

                    # Only count if queryable
                    if is_queryable:
                        count = self.count_records(instance_url, access_token, obj_name)
                        count_str = str(count) if count >= 0 else "Error"
                    else:
                        count_str = "N/A"

                    queryable_str = "✓" if is_queryable else "✗"
                    table.add_row(obj_name, obj_label, count_str, queryable_str)

            console.print(table)
            console.print(f"\n[green]Displayed {len(filtered_objects)} object(s)[/green]\n")

        except Exception as e:
            console.print(f"[bold red]Error: {str(e)}[/bold red]")
            raise


def main():
    """Main entry point"""
    console.print("\n[bold blue]═══════════════════════════════════════════════════════[/bold blue]")
    console.print("[bold blue]       Salesforce Query Tool via AWS Secrets          [/bold blue]")
    console.print("[bold blue]═══════════════════════════════════════════════════════[/bold blue]\n")

    # Create tool instance
    tool = SalesforceQueryTool(
        profile=CONFIG['aws_profile'],
        region=CONFIG['aws_region']
    )

    # Check if we're in search objects mode
    if CONFIG['search_objects_mode']:
        if not CONFIG['secret_path']:
            console.print("[red]Error: 'secret_path' is not set in CONFIG[/red]")
            console.print("[yellow]Please edit the CONFIG section at the top of this file[/yellow]")
            return

        tool.run_search_objects(
            secret_path=CONFIG['secret_path'],
            search_filter=CONFIG['search_objects_filter'],
            always_use_oauth=CONFIG['always_use_oauth']
        )
        return

    # Validate configuration
    if not CONFIG['secret_path']:
        console.print("[red]Error: 'secret_path' is not set in CONFIG[/red]")
        console.print("[yellow]Please edit the CONFIG section at the top of this file[/yellow]")
        return

    if not CONFIG['sobject']:
        console.print("[red]Error: 'sobject' is not set in CONFIG[/red]")
        console.print("[yellow]Please edit the CONFIG section at the top of this file[/yellow]")
        return

    # Run the query
    tool.run_query(
        secret_path=CONFIG['secret_path'],
        sobject=CONFIG['sobject'],
        query_type=CONFIG['query_type'],
        custom_query=CONFIG['custom_query'],
        auto_refresh=CONFIG['auto_refresh_on_expire'],
        always_use_oauth=CONFIG['always_use_oauth']
    )


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        console.print("\n[yellow]Interrupted by user[/yellow]")
    except Exception as e:
        console.print(f"\n[bold red]Fatal error: {str(e)}[/bold red]")